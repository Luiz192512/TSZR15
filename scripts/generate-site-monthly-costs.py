from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
DOCS_DIR = ROOT / "docs"
ACCESS_DATE = date(2026, 4, 28)
OUTPUT_STEM = f"site-monthly-costs-{ACCESS_DATE.isoformat()}"


HEADER_FILL = PatternFill(fill_type="solid", fgColor="0F172A")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="DBEAFE")
GOOD_FILL = PatternFill(fill_type="solid", fgColor="DCFCE7")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FEF3C7")
MUTED_FILL = PatternFill(fill_type="solid", fgColor="F1F5F9")


def autosize(worksheet) -> None:
    for idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 12), 54)


def style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_assumptions(workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "Premissas"
    ws.append(["Campo", "Valor", "Unidade", "Observacao"])
    style_header(ws[1])

    rows = [
        ("Data da analise", ACCESS_DATE.isoformat(), "", "Planilha gerada com precos oficiais consultados em 2026-04-28."),
        ("USD/BRL PTAX fechamento", 4.9700, "BRL por USD", "Banco Central do Brasil, fechamento de 2026-04-27."),
        ("Pedidos por mes", 100, "pedidos", "Valor editavel."),
        ("Ticket medio", 250, "BRL", "Valor editavel."),
        ("GMV mensal", "=B3*B4", "BRL", "Faturamento bruto mensal usado nas taxas variaveis."),
        ("Mix Pix", 0.60, "fracao", "Valor editavel."),
        ("Mix cartao", 0.35, "fracao", "Valor editavel."),
        ("Mix boleto", 0.05, "fracao", "Valor editavel."),
        ("Taxa cartao MP em 30 dias", 0.0399, "fracao", "Checkout online; fonte oficial Mercado Pago."),
        ("Taxa cartao MP na hora", 0.0499, "fracao", "Alternativa para comparacao de fluxo de caixa."),
        ("Taxa Pix MP", 0.0099, "fracao", "Checkout online; fonte oficial Mercado Pago."),
        ("Taxa boleto MP", 3.49, "BRL por boleto", "Checkout online; fonte oficial Mercado Pago."),
        ("Storage R2 usado", 20, "GB-mes", "Exemplo para imagens e modelos 3D; valor editavel."),
        ("Operacoes R2 Class A", 0.10, "milhoes/mes", "Exemplo; valor editavel."),
        ("Operacoes R2 Class B", 2, "milhoes/mes", "Exemplo; valor editavel."),
        ("Emails transacionais por mes", 1500, "emails", "Exemplo; 15 emails por 10 pedidos."),
        ("Caixas Google Workspace", 1, "usuarios", "Exemplo para suporte@ ou contato@."),
        ("Usar UptimeRobot Solo", 1, "0 ou 1", "Troque para 0 se nao quiser custo opcional."),
        ("Usar Resend Pro", 0, "0 ou 1", "0 usa Free; 1 usa Pro."),
        ("Transferencia extra Vercel", 0, "GB/mes", "Acima do que estiver incluido no plano."),
        ("Edge requests extras Vercel", 0, "milhoes/mes", "Acima do que estiver incluido no plano."),
        ("IOF sobre gastos internacionais", 0.035, "fracao", "Usado para estimar custo em cartao em SaaS cobrados em USD."),
    ]

    for row in rows:
        ws.append(list(row))

    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[3].alignment = Alignment(wrap_text=True, vertical="top")

    for idx in range(2, ws.max_row + 1):
        if idx in {3, 4, 5, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22}:
            ws[f"B{idx}"].number_format = "#,##0.00"
        if idx in {7, 8, 9, 10, 11, 12, 23}:
            ws[f"B{idx}"].number_format = "0.00%"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def write_costs(workbook: Workbook) -> None:
    ws = workbook.create_sheet("Custos")
    ws.append(
        [
            "Categoria",
            "Item",
            "Provider",
            "Status",
            "Tipo",
            "Moeda",
            "Preco unitario",
            "Quantidade / formula",
            "Mensal moeda origem",
            "Mensal em BRL",
            "Incluir no total",
            "Notas",
            "Fonte",
        ]
    )
    style_header(ws[1])

    rows = [
        [
            "Dominio",
            "Dominio .com.br",
            "Registro.br",
            "Obrigatorio",
            "Fixo",
            "BRL",
            40,
            "=1/12",
            "=G2*H2",
            "=I2",
            1,
            "Registro anual de dominio brasileiro; mensalizado para orcamento.",
            "https://registro.br/",
        ],
        [
            "Dominio",
            "Dominio .com (alternativa)",
            "Porkbun",
            "Alternativa",
            "Fixo",
            "USD",
            11.08,
            "=1/12",
            "=G3*H3",
            "=I3*Premissas!$B$3",
            0,
            "Alternativa internacional; use se decidir operar com .com em vez de .com.br.",
            "https://porkbun.com/products/domains",
        ],
        [
            "Hospedagem",
            "Frontend / app Next.js - Pro",
            "Vercel",
            "Obrigatorio",
            "Fixo",
            "USD",
            20,
            1,
            "=G4*H4",
            "=I4*Premissas!$B$3",
            1,
            "Plano Pro para ambiente de producao. Hobby existe, mas este orcamento usa baseline profissional.",
            "https://vercel.com/pricing",
        ],
        [
            "Hospedagem",
            "Transferencia extra",
            "Vercel",
            "Sob demanda",
            "Overage",
            "USD",
            0.15,
            "=Premissas!$B$21",
            "=G5*H5",
            "=I5*Premissas!$B$3",
            1,
            "So cobra se passar do incluido no plano.",
            "https://vercel.com/pricing",
        ],
        [
            "Hospedagem",
            "Edge requests extras",
            "Vercel",
            "Sob demanda",
            "Overage",
            "USD",
            2,
            "=Premissas!$B$22",
            "=G6*H6",
            "=I6*Premissas!$B$3",
            1,
            "Preco por 1M de requests acima do incluido no Pro.",
            "https://vercel.com/pricing",
        ],
        [
            "Banco de dados",
            "Postgres gerenciado - Launch",
            "Neon",
            "Obrigatorio",
            "Fixo estimado",
            "USD",
            15,
            1,
            "=G7*H7",
            "=I7*Premissas!$B$3",
            1,
            "Usa o typical spend oficial do plano Launch como baseline do projeto.",
            "https://neon.com/pricing",
        ],
        [
            "Storage de assets",
            "Storage acima do free tier",
            "Cloudflare R2",
            "Sob demanda",
            "Variavel",
            "USD",
            0.015,
            '=MAX(0,Premissas!$B$14-10)',
            "=G8*H8",
            "=I8*Premissas!$B$3",
            1,
            "O free tier inclui 10 GB-mes; este item cobra apenas excedente.",
            "https://developers.cloudflare.com/r2/pricing/",
        ],
        [
            "Storage de assets",
            "Operacoes Class A acima do free tier",
            "Cloudflare R2",
            "Sob demanda",
            "Variavel",
            "USD",
            4.50,
            '=MAX(0,Premissas!$B$15-1)',
            "=G9*H9",
            "=I9*Premissas!$B$3",
            1,
            "Preco por milhao de operacoes acima de 1M/mensal.",
            "https://developers.cloudflare.com/r2/pricing/",
        ],
        [
            "Storage de assets",
            "Operacoes Class B acima do free tier",
            "Cloudflare R2",
            "Sob demanda",
            "Variavel",
            "USD",
            0.36,
            '=MAX(0,Premissas!$B$16-10)',
            "=G10*H10",
            "=I10*Premissas!$B$3",
            1,
            "Preco por milhao de operacoes acima de 10M/mensal.",
            "https://developers.cloudflare.com/r2/pricing/",
        ],
        [
            "Email transacional",
            "Plano Pro",
            "Resend",
            "Opcional",
            "Fixo",
            "USD",
            20,
            "=Premissas!$B$20",
            "=G11*H11",
            "=I11*Premissas!$B$3",
            1,
            "Se B19=0, voce fica no Free ate 3.000 emails/mes; se B19=1, usa Pro com 50.000 emails/mes.",
            "https://resend.com/pricing",
        ],
        [
            "Email corporativo",
            "Business Starter",
            "Google Workspace",
            "Recomendado",
            "Fixo",
            "USD",
            7,
            "=Premissas!$B$18",
            "=G12*H12",
            "=I12*Premissas!$B$3",
            1,
            "Usado para caixa profissional com dominio proprio.",
            "https://workspace.google.com/pricing",
        ],
        [
            "Monitoramento",
            "Plano Solo",
            "UptimeRobot",
            "Opcional",
            "Fixo",
            "USD",
            7,
            "=Premissas!$B$19",
            "=G13*H13",
            "=I13*Premissas!$B$3",
            1,
            "Monitoramento de disponibilidade e expirar de SSL/dominio.",
            "https://uptimerobot.com/pricing/",
        ],
        [
            "Pagamentos",
            "Taxa Pix",
            "Mercado Pago",
            "Obrigatorio",
            "Variavel",
            "BRL",
            "=Premissas!$B$12",
            "=Premissas!$B$6*Premissas!$B$7",
            "=G14*H14",
            "=I14",
            1,
            "Baseado no mix de Pix e na taxa online oficial.",
            "https://www.mercadopago.com.br/blog/quanto-custa-vender-on-line-com-mercado-pago",
        ],
        [
            "Pagamentos",
            "Taxa cartao (recebendo em 30 dias)",
            "Mercado Pago",
            "Obrigatorio",
            "Variavel",
            "BRL",
            "=Premissas!$B$10",
            "=Premissas!$B$6*Premissas!$B$8",
            "=G15*H15",
            "=I15",
            1,
            "Use esta linha como baseline. Se quiser comparar recebimento imediato, veja a linha abaixo.",
            "https://www.mercadopago.com.br/blog/quanto-custa-vender-on-line-com-mercado-pago",
        ],
        [
            "Pagamentos",
            "Delta se trocar cartao para recebimento na hora",
            "Mercado Pago",
            "Cenario",
            "Variavel",
            "BRL",
            "=(Premissas!$B$11-Premissas!$B$10)",
            "=Premissas!$B$6*Premissas!$B$8",
            "=G16*H16",
            "=I16",
            0,
            "Mantido fora do total para servir de comparacao de fluxo de caixa.",
            "https://www.mercadopago.com.br/blog/quanto-custa-vender-on-line-com-mercado-pago",
        ],
        [
            "Pagamentos",
            "Taxa boleto",
            "Mercado Pago",
            "Obrigatorio",
            "Variavel",
            "BRL",
            "=Premissas!$B$13",
            "=Premissas!$B$4*Premissas!$B$9",
            "=G17*H17",
            "=I17",
            1,
            "Custo por boleto emitido no mix informado.",
            "https://www.mercadopago.com.br/blog/quanto-custa-vender-on-line-com-mercado-pago",
        ],
        [
            "Tributo financeiro",
            "IOF sobre SaaS em USD",
            "Cartao / emissor",
            "Sob demanda",
            "Variavel",
            "BRL",
            '=Premissas!$B$23*SUMPRODUCT(($F$2:$F$13="USD")*($K$2:$K$13=1)*($I$2:$I$13))*Premissas!$B$3',
            1,
            "=G18*H18",
            "=I18",
            1,
            "Estimativa do imposto em cobrancas internacionais pagas com cartao brasileiro.",
            "https://www.gov.br/fazenda/pt-br/central-de-conteudo/publicacoes/apresentacoes/2025/Maio/iof-maio-2025.pdf/%40%40download/file",
        ],
        [
            "Infra incluida",
            "SSL/TLS",
            "Vercel / Let's Encrypt",
            "Incluido",
            "Zero",
            "USD",
            0,
            1,
            "=0",
            "=0",
            0,
            "Nao orcado separadamente; tratado como custo zero no stack sugerido.",
            "https://vercel.com/pricing",
        ],
        [
            "Infra incluida",
            "CDN global",
            "Vercel",
            "Incluido",
            "Zero",
            "USD",
            0,
            1,
            "=0",
            "=0",
            0,
            "Incluso no plano; overage aparece nas linhas de transferencia e edge requests.",
            "https://vercel.com/pricing",
        ],
        [
            "Infra incluida",
            "Backups base / time-travel",
            "Neon",
            "Incluido",
            "Zero",
            "USD",
            0,
            1,
            "=0",
            "=0",
            0,
            "Ja contemplado no baseline do Neon Launch.",
            "https://neon.com/pricing",
        ],
        [
            "Gateway",
            "Mensalidade do meio de pagamento",
            "Mercado Pago",
            "Incluido",
            "Zero",
            "BRL",
            0,
            1,
            "=0",
            "=0",
            0,
            "O gateway online modelado aqui nao entra com mensalidade fixa; o custo vem das taxas por venda.",
            "https://www.mercadopago.com.br/developers/pt/support/37740",
        ],
    ]

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    money_cols = {"G", "I", "J"}
    percent_rows = {14, 15, 16}

    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        ws[f"K{row_idx}"].alignment = Alignment(horizontal="center", vertical="top")
        if ws[f"D{row_idx}"].value in {"Obrigatorio", "Recomendado"}:
            for cell in ws[row_idx]:
                cell.fill = GOOD_FILL if ws[f"D{row_idx}"].value == "Obrigatorio" else WARN_FILL
        if ws[f"D{row_idx}"].value == "Incluido":
            for cell in ws[row_idx]:
                cell.fill = MUTED_FILL

        if row_idx in percent_rows:
            ws[f"G{row_idx}"].number_format = "0.00%"
        else:
            ws[f"G{row_idx}"].number_format = "#,##0.00"

        ws[f"I{row_idx}"].number_format = "#,##0.00"
        ws[f"J{row_idx}"].number_format = '"R$" #,##0.00'

    autosize(ws)


def write_summary(workbook: Workbook) -> None:
    ws = workbook.create_sheet("Resumo")
    ws.append(["Resumo", "Formula / valor", "Observacao"])
    style_header(ws[1])

    rows = [
        ("Custos fixos obrigatorios", '=SUMPRODUCT((Custos!$E$2:$E$22<>"Variavel")*(Custos!$E$2:$E$22<>"Overage")*(Custos!$D$2:$D$22="Obrigatorio")*(Custos!$K$2:$K$22=1)*Custos!$J$2:$J$22)', "Sem custos opcionais e sem taxas por venda."),
        ("Custos fixos recomendados/opcionais ligados", '=SUMPRODUCT((Custos!$E$2:$E$22<>"Variavel")*(Custos!$E$2:$E$22<>"Overage")*(Custos!$K$2:$K$22=1)*Custos!$J$2:$J$22)', "Inclui linhas opcionais marcadas em Premissas."),
        ("Custos variaveis de pagamento no cenario atual", '=SUMPRODUCT((Custos!$E$2:$E$22="Variavel")*(Custos!$K$2:$K$22=1)*Custos!$J$2:$J$22)', "Baseado em pedidos, ticket e mix de pagamento das Premissas."),
        ("Overages tecnicos no cenario atual", '=SUMPRODUCT((Custos!$E$2:$E$22="Overage")*(Custos!$K$2:$K$22=1)*Custos!$J$2:$J$22)', "Normalmente zero se o uso ficar no baseline."),
        ("Total mensal operacional", '=B2+B3+B4+B5', "Soma do que entra no orcamento mensal do site."),
        ("Delta se cartao receber na hora", '=Custos!J16', "Custo extra mensal se trocar o cartao de 30 dias para recebimento imediato."),
        ("Total mensal com cartao na hora", '=B6+B7', "Cenario alternativo de caixa."),
        ("Observacao de escopo", "Fora do escopo", "Nao inclui custo do produto, frete do dropshipping, impostos da empresa, devolucoes, chargeback, dev e design."),
    ]

    for row in rows:
        ws.append(list(row))

    for idx in range(2, 8):
        ws[f"B{idx}"].number_format = '"R$" #,##0.00'

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def write_sources(workbook: Workbook) -> None:
    ws = workbook.create_sheet("Fontes")
    ws.append(["Fonte", "URL", "Uso", "Data de acesso"])
    style_header(ws[1])

    rows = [
        ("Registro.br", "https://registro.br/", "Meta description com preco de dominio .br", ACCESS_DATE.isoformat()),
        ("Porkbun", "https://porkbun.com/products/domains", "Preco de entrada para .com", ACCESS_DATE.isoformat()),
        ("Vercel Pricing", "https://vercel.com/pricing", "Plano Pro e overages", ACCESS_DATE.isoformat()),
        ("Neon Pricing", "https://neon.com/pricing", "Typical spend e preco Launch", ACCESS_DATE.isoformat()),
        ("Cloudflare R2 Pricing", "https://developers.cloudflare.com/r2/pricing/", "Storage, operacoes e free tier", ACCESS_DATE.isoformat()),
        ("Resend Pricing", "https://resend.com/pricing", "Free e Pro para email transacional", ACCESS_DATE.isoformat()),
        ("Google Workspace Pricing", "https://workspace.google.com/pricing", "Business Starter por usuario", ACCESS_DATE.isoformat()),
        ("UptimeRobot Pricing", "https://uptimerobot.com/pricing/", "Plano Solo de monitoramento", ACCESS_DATE.isoformat()),
        ("Mercado Pago blog oficial", "https://www.mercadopago.com.br/blog/quanto-custa-vender-on-line-com-mercado-pago", "Taxas online de cartao, Pix e boleto", ACCESS_DATE.isoformat()),
        ("Mercado Pago support", "https://www.mercadopago.com.br/developers/pt/support/37740", "Confirmacao de que o custo varia por meio e prazo", ACCESS_DATE.isoformat()),
        ("BCB PTAX", "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/CotacaoMoedaPeriodo", "Conversao USD/BRL usada na planilha", ACCESS_DATE.isoformat()),
        ("Gov.br Fazenda", "https://www.gov.br/fazenda/pt-br/central-de-conteudo/publicacoes/apresentacoes/2025/Maio/iof-maio-2025.pdf/%40%40download/file", "Aliquota de 3,5% para cartoes internacionais", ACCESS_DATE.isoformat()),
    ]

    for row in rows:
        ws.append(list(row))

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def render_markdown() -> str:
    return "\n".join(
        [
            f"# Custos mensais do site - {ACCESS_DATE.isoformat()}",
            "",
            "## Escopo",
            "",
            "- Inclui custo operacional mensal do site em producao.",
            "- Nao inclui desenvolvimento, design, custo do produto, frete, imposto da empresa ou chargeback.",
            "",
            "## Stack orcada",
            "",
            "- Dominio .com.br em Registro.br.",
            "- App Next.js em Vercel Pro.",
            "- Banco em Neon Launch.",
            "- Assets em Cloudflare R2.",
            "- Pagamentos via Mercado Pago.",
            "- Email transacional em Resend e caixa corporativa em Google Workspace.",
            "",
            "## Leitura rapida",
            "",
            "- A maior parte do custo fixo fica em hospedagem, banco e email corporativo.",
            "- Em operacao real, as taxas do Mercado Pago tendem a ser o maior custo mensal variavel do site.",
            "- Dominio e storage ficam relativamente pequenos perto de pagamentos.",
            "- SaaS internacionais pagos em cartao brasileiro podem gerar IOF; a planilha modela isso separadamente.",
            "",
            "## Arquivos",
            "",
            f"- `docs/{OUTPUT_STEM}.xlsx`",
            f"- `docs/{OUTPUT_STEM}.md`",
            "",
            "## Fontes principais",
            "",
            "- https://registro.br/",
            "- https://porkbun.com/products/domains",
            "- https://vercel.com/pricing",
            "- https://neon.com/pricing",
            "- https://developers.cloudflare.com/r2/pricing/",
            "- https://resend.com/pricing",
            "- https://workspace.google.com/pricing",
            "- https://uptimerobot.com/pricing/",
            "- https://www.mercadopago.com.br/blog/quanto-custa-vender-on-line-com-mercado-pago",
            "- https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/CotacaoMoedaPeriodo",
            "- https://www.gov.br/fazenda/pt-br/central-de-conteudo/publicacoes/apresentacoes/2025/Maio/iof-maio-2025.pdf/%40%40download/file",
            "",
        ]
    )


def main() -> None:
    DOCS_DIR.mkdir(exist_ok=True)

    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    write_assumptions(workbook)
    write_costs(workbook)
    write_summary(workbook)
    write_sources(workbook)

    xlsx_path = DOCS_DIR / f"{OUTPUT_STEM}.xlsx"
    workbook.save(xlsx_path)

    markdown_path = DOCS_DIR / f"{OUTPUT_STEM}.md"
    markdown_path.write_text(render_markdown(), encoding="utf-8")

    print(f"Wrote {xlsx_path}")
    print(f"Wrote {markdown_path}")


if __name__ == "__main__":
    main()
