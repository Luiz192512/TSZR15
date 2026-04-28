from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
DOCS_DIR = ROOT / "docs"
ACCESS_DATE = date(2026, 4, 28)
OUTPUT_STEM = f"database-comparison-{ACCESS_DATE.isoformat()}"


@dataclass(frozen=True)
class Provider:
    name: str
    entry_plan: str
    pricing_model: str
    estimated_monthly_floor_usd: str
    brazil_region: str
    backups_and_recovery: str
    pooling: str
    strengths: str
    caveats: str
    best_when: str
    notes: str
    cost_score: int
    brazil_score: int
    ops_score: int
    feature_score: int
    growth_score: int

    @property
    def total_score(self) -> int:
        weighted = (
            self.cost_score * 0.35
            + self.brazil_score * 0.20
            + self.ops_score * 0.20
            + self.feature_score * 0.15
            + self.growth_score * 0.10
        )
        return round(weighted * 20)


PROFILE_ROWS = [
    ("Projeto", "Loja customizada de acessorios para Yamaha R15 em Next.js"),
    ("Tipo de banco", "Relacional, compatível com PostgreSQL"),
    ("Carga inicial esperada", "Baixa a moderada; MVP com catalogo, configuracoes salvas, pedidos e webhooks"),
    ("Publico principal", "Brasil, com preferencia por regiao Sao Paulo quando existir"),
    ("Prioridade", "Custo-beneficio, baixa operacao e backups nativos"),
    ("Requisito tecnico", "Boa compatibilidade com serverless/Next.js e idealmente connection pooling"),
    ("Nao e prioridade no MVP", "Data warehouse, multi-cloud, analiticos pesados e compliance enterprise"),
    (
        "Escopo da comparacao",
        "Comparacao entre bancos gerenciados relacionais que fazem sentido para o projeto. "
        "Nao inclui servicos NoSQL, bancos edge baseados em SQLite, ou ofertas white-label equivalentes.",
    ),
]


PROVIDERS = [
    Provider(
        name="Neon",
        entry_plan="Free $0; Launch usage-based",
        pricing_model="Serverless Postgres com scale-to-zero; $0.106/CU-h + $0.35/GB-mes no Launch",
        estimated_monthly_floor_usd="$15 tipico/mes (oficial) para carga intermitente com 1 GB",
        brazil_region="Sim, AWS Sao Paulo (aws-sa-east-1)",
        backups_and_recovery="Time travel 6h no Free e 7 dias no Launch",
        pooling="Sim, pooling integrado com pgBouncer; ate 10.000 conexoes",
        strengths="Melhor custo-beneficio DB-only, DX muito boa, branching, baixa operacao",
        caveats="Custo e variavel por uso; nao traz auth/storage completos como plataforma",
        best_when="Voce quer banco gerenciado enxuto para um app Next.js customizado",
        notes="Minha recomendacao principal para este projeto se quisermos banco em nuvem sem pagar stack extra.",
        cost_score=5,
        brazil_score=5,
        ops_score=5,
        feature_score=4,
        growth_score=4,
    ),
    Provider(
        name="Supabase",
        entry_plan="Pro $25/mes",
        pricing_model="Assinatura + compute por hora; creditos cobrem um projeto Micro (~$10/mes)",
        estimated_monthly_floor_usd="$25/mes para um projeto Micro no Pro",
        brazil_region="Sim, Sao Paulo (sa-east-1)",
        backups_and_recovery="Backups diarios; 7 dias no Pro; PITR e cobrado a parte",
        pooling="Sim, ecossistema Postgres completo para app/backend",
        strengths="Auth, storage, API e dashboard prontos; acelera muito o backend",
        caveats="Piso de custo maior se voce precisa apenas do banco",
        best_when="Voce quer subir banco + auth + storage + API no mesmo fornecedor",
        notes="Se a ideia mudar para plataforma completa, e a alternativa mais forte ao Neon.",
        cost_score=3,
        brazil_score=5,
        ops_score=5,
        feature_score=5,
        growth_score=4,
    ),
    Provider(
        name="Aiven for PostgreSQL",
        entry_plan="Free $0; Developer a partir de $5/mes",
        pricing_model="Cobranca por hora; custos all-inclusive por plano",
        estimated_monthly_floor_usd="$5/mes no Developer; Free sem limite de tempo",
        brazil_region="Sim, AWS Sao Paulo (aws-sa-east-1)",
        backups_and_recovery="Backups incluidos; HA so em Business/Premium",
        pooling="Pooling documentado para Startup/Business/Premium",
        strengths="Postgres gerenciado bem maduro, regiao Brasil, boa reputacao enterprise",
        caveats="Recursos de producao e pooling nao aparecem no tier barato",
        best_when="Voce quer uma alternativa mais tradicional ao Neon, mantendo Brasil como opcao",
        notes="Bom plano B com Brasil; perde pontos no custo real de producao comparado ao Neon.",
        cost_score=4,
        brazil_score=5,
        ops_score=4,
        feature_score=3,
        growth_score=4,
    ),
    Provider(
        name="Railway PostgreSQL",
        entry_plan="Hobby com minimo de $5/mes",
        pricing_model="Uso por segundo; memoria, CPU, volume e egress cobrados separadamente",
        estimated_monthly_floor_usd="$5 minimo/mes, podendo subir conforme uso real",
        brazil_region="Nao; regioes publicas atuais: Virginia, California, Amsterdam e Singapore",
        backups_and_recovery="Depende da forma de operacao no projeto; plataforma foca mais em deploy do que em DB puro",
        pooling="Nao destacado como recurso nativo de banco gerenciado no material principal",
        strengths="Otimo quando app e banco ficam na mesma plataforma e voce quer deploy simples",
        caveats="Sem Sao Paulo; custo real e menos previsivel que Render/Aiven",
        best_when="Voce quer hospedar app e banco juntos na Railway",
        notes="Para DB-only eu priorizaria Neon/Aiven antes; para stack toda na Railway faz sentido.",
        cost_score=4,
        brazil_score=2,
        ops_score=4,
        feature_score=4,
        growth_score=3,
    ),
    Provider(
        name="Render Postgres",
        entry_plan="Basic-256mb $6/mes; Basic-1gb $19/mes",
        pricing_model="Preco fixo por instancia + $0.30/GB de storage expansivel",
        estimated_monthly_floor_usd="$6/mes de entrada; $19/mes fica mais realista para producao leve",
        brazil_region="Nao; regioes publicas: Oregon, Ohio, Virginia, Frankfurt e Singapore",
        backups_and_recovery="PITR em bases pagas; 3 dias no Hobby e 7 dias em planos maiores",
        pooling="Suportado via PgBouncer separado",
        strengths="Preco previsivel, simples, boa experiencia para stacks pequenas",
        caveats="Sem Sao Paulo; HA so em specs elegiveis",
        best_when="Voce quer previsibilidade e simplicidade acima de latencia Brasil",
        notes="Boa opcao simples, mas perde pontos para nosso publico brasileiro.",
        cost_score=4,
        brazil_score=2,
        ops_score=4,
        feature_score=3,
        growth_score=3,
    ),
    Provider(
        name="DigitalOcean Managed PostgreSQL",
        entry_plan="1 GiB / 1 vCPU / 10-30 GiB a $15.15/mes",
        pricing_model="Preco fixo por node + storage adicional a $0.215/GiB-mes",
        estimated_monthly_floor_usd="$15.15/mes",
        brazil_region="Nao evidenciado nas fontes revisadas para managed DB em Sao Paulo",
        backups_and_recovery="Backups diarios, PITR e standby/documentacao de failover",
        pooling="Nao destacado como diferencial no material principal revisado",
        strengths="Preco previsivel, managed DB conhecido, boa operacao padrao",
        caveats="Floor maior que Neon/Render/Railway e sem vantagem clara de Brasil nas fontes vistas",
        best_when="Voce quer flat pricing e ja usa DigitalOcean",
        notes="Nao e ruim; apenas nao aparece como melhor custo-beneficio para o nosso perfil.",
        cost_score=3,
        brazil_score=1,
        ops_score=4,
        feature_score=4,
        growth_score=4,
    ),
    Provider(
        name="AWS RDS for PostgreSQL",
        entry_plan="Free Tier por 12 meses; depois on-demand",
        pricing_model="Instancia + storage + backups/IO conforme configuracao",
        estimated_monthly_floor_usd="~$29.68/mes em Sao Paulo para db.t4g.micro + 20 GB gp3",
        brazil_region="Sim, South America (Sao Paulo / sa-east-1)",
        backups_and_recovery="Backups automatizados e PITR de ate 35 dias",
        pooling="Nao nativo como recurso gerenciado simples; normalmente exige camada adicional",
        strengths="Muita maturidade, ecossistema enorme, excelente opcao de crescimento",
        caveats="Mais complexo para operar e mais caro no inicio",
        best_when="Voce aceita mais operacao em troca de padroes AWS e previsao de escalar forte",
        notes="Bom tecnicamente, mas eu nao comecaria aqui para este MVP.",
        cost_score=2,
        brazil_score=5,
        ops_score=2,
        feature_score=5,
        growth_score=5,
    ),
    Provider(
        name="Azure Database for PostgreSQL",
        entry_plan="B1ms (1 vCore / 2 GiB) com storage separado",
        pricing_model="Compute por hora + storage Premium SSD por GiB/mes",
        estimated_monthly_floor_usd="~$33.03/mes em Brazil South usando B1ms + 32 GiB de storage",
        brazil_region="Sim, Brazil South",
        backups_and_recovery="Backup storage ate 100% do provisionado sem custo adicional",
        pooling="Nao destacado como diferencial simples no material revisado",
        strengths="Boa presenca no Brasil e integracao forte com Azure",
        caveats="Preco e operacao menos atraentes para nosso MVP que Neon/Supabase",
        best_when="Voce ja esta no ecossistema Azure e quer tudo centralizado",
        notes="Valor mensal acima das melhores opcoes para este caso.",
        cost_score=2,
        brazil_score=5,
        ops_score=2,
        feature_score=4,
        growth_score=5,
    ),
    Provider(
        name="Google Cloud SQL for PostgreSQL",
        entry_plan="Sem plano fixo; CPU + memoria + storage cobrados por uso",
        pricing_model="vCPU/h + GiB/h + storage; calculadora por regiao",
        estimated_monthly_floor_usd="~$51.99/mes em exemplo Iowa com 1 vCPU + 3.75 GiB + 10 GB; Sao Paulo varia",
        brazil_region="Sim, Sao Paulo (southamerica-east1)",
        backups_and_recovery="Modelo completo de Cloud SQL com backups e opcoes HA",
        pooling="Nao destacado como principal diferencial no material revisado",
        strengths="Forte integracao GCP e bom caminho para workloads maiores",
        caveats="Comparacao mais dificil, piso inicial mais alto e menos amigavel para MVP economico",
        best_when="Voce ja esta no GCP e precisa alinhar o banco ao restante da infra",
        notes="Para nosso perfil, o custo de entrada nao ficou competitivo.",
        cost_score=1,
        brazil_score=5,
        ops_score=2,
        feature_score=4,
        growth_score=5,
    ),
]


RANKING_ROWS = [
    ("Melhor custo-beneficio DB-only", "Neon", "Banco gerenciado com Brasil, pooling e custo inicial muito bom."),
    ("Melhor opcao all-in-one", "Supabase", "Vale se voce quiser banco + auth + storage + APIs prontas."),
    ("Melhor alternativa com Brasil fora Neon/Supabase", "Aiven", "Tem Sao Paulo e tier barato, mas perde em recursos nos tiers de entrada."),
    ("Melhor se app e banco ficarem juntos", "Railway", "A plataforma inteira fica simples, mesmo sem regiao Brasil."),
]


SOURCES = [
    ("Neon", "Pricing", "https://neon.com/pricing", "Launch usage-based, typical spend, storage e pooling"),
    ("Neon", "Regions", "https://neon.com/docs/conceptual-guides/regions", "Disponibilidade em AWS Sao Paulo"),
    ("Supabase", "Compute pricing", "https://supabase.com/docs/guides/platform/manage-your-usage/compute", "Pro $25 + creditos de compute e valores Micro"),
    ("Supabase", "Regions", "https://supabase.com/docs/guides/platform/regions", "Disponibilidade em Sao Paulo"),
    ("Supabase", "Backups", "https://supabase.com/docs/guides/platform/backups", "Backups diarios e PITR"),
    ("Railway", "Pricing", "https://railway.com/pricing", "Minimos de Hobby/Pro e cobranca por recursos"),
    ("Railway", "Regions", "https://docs.railway.com/reference/regions", "Regioes publicas disponiveis"),
    ("Render", "Pricing", "https://render.com/pricing", "Precos das instancias Postgres e storage"),
    ("Render", "Regions", "https://render.com/docs/regions", "Lista de regioes publicas"),
    ("Render", "Backups", "https://render.com/docs/postgresql-backups", "Janela de PITR por workspace"),
    ("Aiven", "Pricing", "https://aiven.io/pricing", "Free tier e estrutura de planos"),
    ("Aiven", "Developer tier", "https://aiven.io/blog/new-developer-tier-for-aiven-for-postgres", "Developer starting at $5"),
    ("Aiven", "Regions", "https://aiven.io/docs/platform/reference/list_of_clouds", "Disponibilidade em aws-sa-east-1"),
    ("Aiven", "Service pricing", "https://aiven.io/docs/platform/concepts/service-pricing", "Cobranca por hora e custos all-inclusive"),
    ("Aiven", "Connection limits / pooling", "https://aiven.io/docs/products/postgresql/reference/pg-connection-limits", "Pooling nas faixas superiores"),
    ("DigitalOcean", "Managed DB pricing", "https://www.digitalocean.com/pricing/managed-databases", "Preco base PostgreSQL"),
    ("DigitalOcean", "Managed DB docs", "https://docs.digitalocean.com/products/databases/index.html", "Backups, PITR e HA"),
    ("AWS", "RDS PostgreSQL pricing", "https://aws.amazon.com/rds/postgresql/pricing/", "Free tier e modelo de cobranca"),
    ("AWS", "RDS regions", "https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/Concepts.RegionsAndAvailabilityZones.html", "Sao Paulo disponivel"),
    ("Google Cloud", "Cloud SQL pricing", "https://cloud.google.com/sql/pricing", "vCPU, memoria, storage e regiao Sao Paulo"),
    ("Azure", "Azure PostgreSQL pricing", "https://azure.microsoft.com/en-us/pricing/details/postgresql/flexible-server/", "B1ms em Brazil South e storage"),
    ("Azure", "Azure PostgreSQL regions overview", "https://learn.microsoft.com/en-us/azure/postgresql/flexible-server/overview", "Brazil South suportado"),
]


HEADER_FILL = PatternFill(fill_type="solid", fgColor="0F172A")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="DCEBFF")
GOOD_FILL = PatternFill(fill_type="solid", fgColor="DCFCE7")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FEF3C7")


def autosize(worksheet) -> None:
    for idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 14), 52)


def style_header(row: Iterable) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_profile_sheet(workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "Perfil"
    ws.append(["Campo", "Valor"])
    style_header(ws[1])

    for field, value in PROFILE_ROWS:
        ws.append([field, value])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def build_comparison_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("Comparativo")
    ws.append(
        [
            "Provider",
            "Plano de entrada",
            "Modelo de cobranca",
            "Piso mensal estimado (USD)",
            "Regiao Brasil",
            "Backups e recovery",
            "Pooling",
            "Forcas",
            "Caveats",
            "Melhor quando",
            "Notas",
            "Custo (1-5)",
            "Brasil (1-5)",
            "Operacao (1-5)",
            "Fit de recursos (1-5)",
            "Crescimento (1-5)",
            "Score total / 100",
        ]
    )
    style_header(ws[1])

    for provider in sorted(PROVIDERS, key=lambda item: item.total_score, reverse=True):
        ws.append(
            [
                provider.name,
                provider.entry_plan,
                provider.pricing_model,
                provider.estimated_monthly_floor_usd,
                provider.brazil_region,
                provider.backups_and_recovery,
                provider.pooling,
                provider.strengths,
                provider.caveats,
                provider.best_when,
                provider.notes,
                provider.cost_score,
                provider.brazil_score,
                provider.ops_score,
                provider.feature_score,
                provider.growth_score,
                provider.total_score,
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        total_cell = row[16]
        if total_cell.value >= 85:
            total_cell.fill = GOOD_FILL
        elif total_cell.value >= 70:
            total_cell.fill = WARN_FILL

    autosize(ws)


def build_ranking_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("Ranking")
    ws.append(["Cenario", "Recomendado", "Por que"])
    style_header(ws[1])

    for row in RANKING_ROWS:
        ws.append(list(row))

    ws.append([])
    ws.append(["Ranking geral", "Score", "Resumo"])
    for cell in ws[ws.max_row]:
        cell.fill = SECTION_FILL
        cell.font = Font(bold=True)

    for provider in sorted(PROVIDERS, key=lambda item: item.total_score, reverse=True):
        ws.append(
            [
                provider.name,
                provider.total_score,
                f"{provider.best_when} Caveat principal: {provider.caveats}",
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C5"
    autosize(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_sources_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("Fontes")
    ws.append(["Provider", "Fonte", "URL", "Uso", "Acesso"])
    style_header(ws[1])

    for provider, label, url, note in SOURCES:
        ws.append([provider, label, url, note, ACCESS_DATE.isoformat()])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def render_markdown() -> str:
    ranking = sorted(PROVIDERS, key=lambda item: item.total_score, reverse=True)
    lines = [
        f"# Comparativo de Bancos Gerenciados - {ACCESS_DATE.isoformat()}",
        "",
        "## Perfil considerado",
        "",
    ]

    for field, value in PROFILE_ROWS:
        lines.append(f"- **{field}:** {value}")

    lines.extend(
        [
            "",
            "## Ranking resumido",
            "",
        ]
    )

    for idx, provider in enumerate(ranking[:5], start=1):
        lines.append(
            f"{idx}. **{provider.name}** - score {provider.total_score}/100. "
            f"{provider.notes}"
        )

    lines.extend(
        [
            "",
            "## Recomendacao objetiva",
            "",
            "- **Primeira escolha:** Neon, se quisermos apenas o banco com o melhor custo-beneficio.",
            "- **Segunda escolha:** Supabase, se quisermos ganhar auth/storage/API no mesmo fornecedor.",
            "- **Plano B com Brasil:** Aiven, se quisermos um managed Postgres mais tradicional em Sao Paulo.",
            "",
            "## Observacoes de metodologia",
            "",
            "- Os valores foram coletados de fontes oficiais em 2026-04-28.",
            "- Alguns provedores usam precificacao totalmente dinamica. Nesses casos, a planilha marca o calculo como estimativa e explica a base usada.",
            "- O comparativo foca em bancos relacionais gerenciados que fazem sentido para este projeto. Nao inclui NoSQL, SQLite edge DBs ou ofertas equivalentes white-label.",
            "",
            "## Fontes principais",
            "",
        ]
    )

    for provider, label, url, _note in SOURCES:
        lines.append(f"- **{provider} / {label}:** {url}")

    return "\n".join(lines) + "\n"


def main() -> None:
    DOCS_DIR.mkdir(exist_ok=True)

    workbook = Workbook()
    build_profile_sheet(workbook)
    build_comparison_sheet(workbook)
    build_ranking_sheet(workbook)
    build_sources_sheet(workbook)

    xlsx_path = DOCS_DIR / f"{OUTPUT_STEM}.xlsx"
    workbook.save(xlsx_path)

    markdown_path = DOCS_DIR / f"{OUTPUT_STEM}.md"
    markdown_path.write_text(render_markdown(), encoding="utf-8")

    print(f"Wrote {xlsx_path}")
    print(f"Wrote {markdown_path}")


if __name__ == "__main__":
    main()
